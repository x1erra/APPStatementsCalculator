import streamlit as st
from io import BytesIO
from io import StringIO
import base64
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

try:
    from PIL import Image as PILImage
except ImportError:  # pragma: no cover
    PILImage = None


st.set_page_config(
    page_title="Portfolio Composition Calculator",
    page_icon="📊",
    layout="wide",
)


MANUAL_TEXT = """Portfolio Composition Calculator Manual

This manual explains exactly how to produce the two sections of the Portfolio Composition report:

- **Portfolio Composition** (high-level table + donut chart)  
- **Portfolio Breakdown** (detailed SAA vs TAA table + stacked bar chart)

Both sections are created from the same inputs and follow the same process. The only difference is the mapping rule used for each section.

#### 1. Inputs Required

You need exactly these files:

1. **Client holdings file** (Excel)  
   - Must contain the **IPS** sheet with these columns:  
     `Fund Code`, `Fund Description`, `Total MV (CAD)`, `mandate_code`, `saa_taa` (must be “SAA” or “TAA”).

2. **Support files** (exactly the files the client holds — typically 5 files)  
   - 1 TAA file (e.g. 25010.XLSX or 25011.XLSX or 25016.XLSX)  
   - 4 SAA files (e.g. 25000.XLSX, 25004.XLSX, 25006.XLSX, 25008.XLSX)  
   These come from the monthly ZIP folders (SAA/, Pool/, or CC/). Upload only the files whose `mandate_code` appears in the IPS sheet.

3. **Formulas sheet** (copy this from the practice test file)  
   - Contains the Asset Class Detection table, the Funds/Alternatives Detection table, and the long nested IF formula used for the Portfolio Breakdown section.

#### 2. Step-by-Step Process

**Step 1: Extract Client Holdings**  
Open the client holdings file.  
Copy all rows from the IPS sheet that have a `Total MV (CAD)` value into a new sheet called **Holdings**. Keep the columns `Fund Code`, `Fund Description`, `Total MV (CAD)`, `mandate_code`, and `saa_taa`.

**Step 2: Gather Support Files**  
For each row in the Holdings sheet, locate the matching support file using the `mandate_code` value.  
Open each support file and locate the main data block that begins with:  
“Composite Components        Excluded: Multiple Securities”  
the report date shown in that file  
“Port. Weight”

**Step 3: Build the COMP Sheet**  
Create a new sheet named **COMP**.  
For each support file:  
- Copy the entire Port. Weight table (from “Non-Composite” down to the last data row).  
- Paste it into the COMP sheet.  
- Leave 3–4 blank rows between each fund block.  
- Label each block clearly (e.g. “Fixed Income Managed Class – 25000 – SAA”).

**Step 4: Apply Mapping Rules**

**For Portfolio Composition (High-Level – 5 classes)**  
Add two new columns to every row in the COMP sheet:  
- Column: **AMA Group**  
- Column: **Final MV (CAD)**

Apply the two lookup tables from the Formulas sheet in this order:  
1. **Asset Class Detection** (first)  
2. **Funds/Alternatives Detection** (second — overrides the first when it matches)

Resulting high-level groups:  
- Income  
- Equity  
- Cash  
- Other  
- Private Alternatives

For every row calculate:  
**Final MV (CAD) = Client’s Total MV (CAD) × (Port. Weight ÷ 100)**

Sum all Final MV values by AMA Group.

**For Portfolio Breakdown (Detailed – 7 classes + SAA/TAA split)**  
Use the exact nested IF formula provided on the Formulas sheet (the long OR/IF formula).  
This formula produces one of these exact strings:  
- “Fixed Income” → becomes Income row  
- “Equity - Canadian Equities”  
- “Equity - International Equities”  
- “Equity - US Equities”  
- “Cash”  
- “Other”  
- “Alternatives”

For every row:  
- If the parent fund’s `saa_taa` flag (from Holdings sheet) is **SAA**, add the weighted amount to the **Strategic Asset Allocation %** column.  
- If the parent fund’s `saa_taa` flag is **TAA**, add the weighted amount to the **Tactical Asset Allocation %** column.  
- Portfolio % = Strategic % + Tactical %.

**Step 5: Create the Two Final Output Sheets**

**Portfolio Composition**  
Create a sheet with this exact layout:

| Asset Classes          | Portfolio Market Value (CDN) | Actively Managed Market Value (CDN) | % of Portfolio (CDN) |
|------------------------|------------------------------|-------------------------------------|----------------------|
| Income                 | [sum]                        | [same]                              | [calculated]         |
| Equity                 | [sum]                        | [same]                              | [calculated]         |
| Cash                   | [sum]                        | [same]                              | [calculated]         |
| Other                  | [sum]                        | [same]                              | [calculated]         |
| Private Alternatives   | [sum]                        | [same]                              | [calculated]         |
| **Total Market Value of Asset Classes** | [grand total] | [same] | 100.00% |

Add a donut chart with percentage labels.

**Portfolio Breakdown**  
Create a sheet with this exact layout:

| Actively Managed Asset Classes | Strategic Asset Allocation % | Tactical Asset Allocation % | Portfolio % |
|--------------------------------|------------------------------|-----------------------------|-------------|
| Income                         | [SAA sum]                    | [TAA sum]                   | [total]     |
| International Equity           | [SAA sum]                    | [TAA sum]                   | [total]     |
| US Equity                      | [SAA sum]                    | [TAA sum]                   | [total]     |
| Canadian Equity                | [SAA sum]                    | [TAA sum]                   | [total]     |
| Cash                           | [SAA sum]                    | [TAA sum]                   | [total]     |
| Other                          | [SAA sum]                    | [TAA sum]                   | [total]     |
| Alternatives                   | [SAA sum]                    | [TAA sum]                   | [total]     |
| **Total of Actively Managed Assets** | [sum of SAA] | [sum of TAA] | 100.00% |

Add a stacked bar chart (dark blue = Strategic, gray = Tactical).  
Add the following footer notes:  
“Asset Allocation sector weight is reported as of the uploaded support-file period.”  
“The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.”

#### 6. Final Validation Checklist

- Total Market Value in both sections must equal the sum of all `Total MV (CAD)` in the Holdings sheet.  
- In Portfolio Breakdown: Strategic % + Tactical % = Portfolio % for every row.  
- Equity in Portfolio Composition must equal the sum of International Equity + US Equity + Canadian Equity in Portfolio Breakdown.  
- Private Alternatives and Alternatives rows must only contain values from the exact fund names listed in the Formulas sheet.  
- All rows must be mapped (no blanks)."""


COMPOSITION_GROUP_ORDER = [
    "Income",
    "Equity",
    "Cash",
    "Other",
    "Private Alternatives",
]

BREAKDOWN_GROUP_ORDER = [
    "Income",
    "International Equity",
    "US Equity",
    "Canadian Equity",
    "Cash",
    "Other",
    "Alternatives",
]

ASSET_CLASS_DETECTION = {
    "EQUITY": "Equity",
    "CURRENCY FORWARDS": "Other",
    "FIXED INCOME": "Income",
    "CASH": "Cash",
    "DERIVATIVES": "Other",
    "CASH & EQUIVALENTS": "Cash",
    "PREFERRED": "Other",
    "FDS OUTLIER": "Other",
    "[CASH]": "Cash",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND": "Liquid Alt",
}

FUNDS_ALTERNATIVES_DETECTION = {
    "ALATE I LP, RESTRICTED": "Private Alt",
    "AVENUE EUROPE SPECIAL SITUATIONS FUND V (U.S.), L.P.": "Private Alt",
    "AXIA U.S. GROCERY NET LEASE FUND I LP, RESTRICTED": "Private Alt",
    "CI ADAMS STREET GLOBAL PRIVATE MARKETS FUND (CLASS I)": "Private Alt",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND I": "Other",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND (ETF C$ SERIES)": "Other",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND": "Liquid Alt",
    "CI PM GROWTH FUND BL LP (SERIES I)": "Private Alt",
    "CI PRIVATE MARKETS GROWTH FUND I": "Private Alt",
    "CI PRIVATE MARKETS INCOME FUND (SERIES I)": "Other",
    "CI PRIVATE MARKETS GROWTH FUND - SERIES I INSTALLMENT RECEIPT": "Other",
    "CI PRIVATE MARKETS INCOME FUND - SERIES I INSTALLMENT RECEIPT": "Other",
    "HARBOURVEST ADELAIDE FEEDER E LP": "Private Alt",
    "HARBOURVEST ADELAIDE FEEDER F LP": "Private Alt",
    "HARBOURVEST ADELAIDE FEEDER G LP": "Private Alt",
    "HARBOURVEST INFRASTRUCTURE INCOME CAYMAN PARALLEL PARTNERSHIP L.": "Private Alt",
    "INSTITUTIONAL FIDUCIARY TR MONEY MKT PTF": "Equity",
    "INVESCO PREMIER US GOV MONEY PTF": "Other",
    "MONARCH CAPITAL PARTNERS OFFSHORE VI LP": "Private Alt",
    "MSILF PRIME PORTFOLIO-INST": "Equity",
    "T.RX CAPITAL FUND I, LP.": "Private Alt",
    "WHITEHORSE LIQUIDITY PARTNERS V LP": "Private Alt",
}

BREAKDOWN_ALTERNATIVES = {
    "ALATE I LP, RESTRICTED",
    "AVENUE EUROPE SPECIAL SITUATIONS FUND V (U.S.), L.P.",
    "AXIA U.S. GROCERY NET LEASE FUND I LP, RESTRICTED",
    "CI ADAMS STREET GLOBAL PRIVATE MARKETS FUND (CLASS I)",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND",
    "CI PM GROWTH FUND BL LP (SERIES I)",
    "CI PRIVATE MARKETS GROWTH FUND I",
    "DEMOPOLIS EQUITY PARTNERS FUND I, L.P.",
    "HARBOURVEST ADELAIDE FEEDER E LP",
    "HARBOURVEST ADELAIDE FEEDER F LP",
    "HARBOURVEST ADELAIDE FEEDER G LP",
    "HARBOURVEST INFRASTRUCTURE INCOME CAYMAN PARALLEL PARTNERSHIP L.",
    "MONARCH CAPITAL PARTNERS OFFSHORE VI LP",
    "T.RX CAPITAL FUND I, LP.",
    "WHITEHORSE LIQUIDITY PARTNERS V LP",
}

BREAKDOWN_CASH = {
    "CASH & EQUIVALENTS",
    "INSTITUTIONAL FIDUCIARY TR MONEY MKT PTF",
    "MSILF PRIME PORTFOLIO-INST",
    "[CASH]",
}

BREAKDOWN_OTHER = {
    "PREFERRED",
    "CURRENCY FORWARDS",
    "DERIVATIVES",
    "FDS OUTLIER",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND I",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND (ETF C$ SERIES)",
    "CI PRIVATE MARKETS INCOME FUND (SERIES I)",
    "CI PRIVATE MARKETS GROWTH FUND - SERIES I INSTALLMENT RECEIPT",
    "CI PRIVATE MARKETS INCOME FUND - SERIES I INSTALLMENT RECEIPT",
    "INVESCO PREMIER US GOV MONEY PTF",
}

BREAKDOWN_DIRECT = {
    "EQUITY - CANADIAN EQUITIES": "Canadian Equity",
    "EQUITY - INTERNATIONAL EQUITIES": "International Equity",
    "EQUITY - US EQUITIES": "US Equity",
    "FIXED INCOME": "Income",
}

PRIMARY_BLUE = "#25356B"
SECONDARY_GRAY = "#9C9CA6"
MANUAL_HOLDINGS_COLUMNS = ["Fund Code", "Fund Description", "Total MV (CAD)", "saa_taa"]
DEFAULT_HOLDINGS_INPUT = pd.DataFrame(
    [
        {
            "Fund Code": "",
            "Fund Description": "",
            "Total MV (CAD)": None,
            "saa_taa": "SAA",
        }
    ]
)
DRAFT_PATH = Path("/Users/stevendamato/Projects/APPStatements/.portfolio_composition_draft.json")
HOLDINGS_TEXT_TEMPLATE = "Fund Code\tFund Description\tTotal MV (CAD)\tsaa_taa\n"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value: object) -> str:
    return normalize_text(value).upper()


def normalize_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_key(value))


def coerce_number_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "-": pd.NA})
    return pd.to_numeric(cleaned, errors="coerce")


def extract_mandate_code(value: object) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = normalize_text(value)
    match = re.search(r"(\d{5})", text)
    return match.group(1) if match else None


def get_excel_engine(filename: str) -> Optional[str]:
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return "openpyxl"
    if lower.endswith(".xls"):
        return "xlrd"
    return None


@st.cache_data(show_spinner=False)
def read_excel_sheet(file_bytes: bytes, filename: str, sheet_name: object, header=None) -> pd.DataFrame:
    engine = get_excel_engine(filename)
    if engine is None:
        raise ValueError(f"Unsupported Excel file type for `{filename}`.")
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header, engine=engine)
    except ImportError as exc:
        if engine == "openpyxl":
            raise ImportError(
                "Reading .xlsx files requires the `openpyxl` package. Install it with `pip install openpyxl`."
            ) from exc
        if engine == "xlrd":
            raise ImportError(
                "Reading .xls files requires the optional `xlrd` package. Convert the file to .xlsx or install xlrd."
            ) from exc
        raise


@st.cache_data(show_spinner=False)
def list_excel_sheets(file_bytes: bytes, filename: str) -> List[str]:
    engine = get_excel_engine(filename)
    if engine is None:
        raise ValueError(f"Unsupported Excel file type for `{filename}`.")
    with pd.ExcelFile(BytesIO(file_bytes), engine=engine) as workbook:
        return workbook.sheet_names


def find_header_row(raw_df: pd.DataFrame) -> Optional[int]:
    required = {"FUNDCODE", "FUNDDESCRIPTION", "TOTALMVCAD", "MANDATECODE"}
    for idx, row in raw_df.iterrows():
        row_values = {normalize_header(val) for val in row.tolist() if normalize_text(val)}
        if required.issubset(row_values):
            return idx
    return None


def infer_saa_taa_column(df: pd.DataFrame) -> Optional[str]:
    exact = [col for col in df.columns if normalize_header(col) == "SAATAA"]
    if exact:
        return exact[0]

    for col in df.columns:
        values = (
            df[col]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .str.upper()
        )
        if not values.empty and set(values.unique()).issubset({"SAA", "TAA"}):
            return col
    return None


def parse_holdings_file(file_bytes: bytes, filename: str) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    raw_df = read_excel_sheet(file_bytes, filename, sheet_name=0, header=None)
    sheet_names = list_excel_sheets(file_bytes, filename)
    ips_sheet = next((name for name in sheet_names if normalize_key(name) == "IPS"), None)
    if ips_sheet is None:
        raise ValueError("The client holdings file must contain an `IPS` sheet.")

    raw_df = read_excel_sheet(file_bytes, filename, sheet_name=ips_sheet, header=None)
    header_row = find_header_row(raw_df)
    if header_row is None:
        raise ValueError(
            "Unable to locate the IPS header row. Expected columns include `Fund Code`, `Fund Description`, `Total MV (CAD)`, and `mandate_code`."
        )

    header_values = [normalize_text(value) for value in raw_df.iloc[header_row].tolist()]
    data = raw_df.iloc[header_row + 1 :].copy()
    data.columns = [
        value if value else f"Unnamed_{idx}"
        for idx, value in enumerate(header_values)
    ]
    data = data.dropna(how="all").reset_index(drop=True)

    column_map: Dict[str, str] = {}
    for col in data.columns:
        normalized = normalize_header(col)
        if normalized == "FUNDCODE":
            column_map[col] = "Fund Code"
        elif normalized == "FUNDDESCRIPTION":
            column_map[col] = "Fund Description"
        elif normalized == "TOTALMVCAD":
            column_map[col] = "Total MV (CAD)"
        elif normalized == "MANDATECODE":
            column_map[col] = "mandate_code"
        elif normalized == "SAATAA":
            column_map[col] = "saa_taa"

    data = data.rename(columns=column_map)

    saa_taa_column = "saa_taa" if "saa_taa" in data.columns else None
    if saa_taa_column is None:
        inferred = infer_saa_taa_column(data)
        if inferred is None:
            raise ValueError("Unable to locate the `saa_taa` column in the IPS sheet.")
        data = data.rename(columns={inferred: "saa_taa"})

    required_columns = ["Fund Code", "Fund Description", "Total MV (CAD)", "mandate_code", "saa_taa"]
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise ValueError(f"Missing required IPS columns: {', '.join(missing_columns)}")

    holdings = data[required_columns].copy()
    holdings["Total MV (CAD)"] = coerce_number_series(holdings["Total MV (CAD)"])
    holdings["mandate_code"] = holdings["mandate_code"].apply(extract_mandate_code)
    holdings["saa_taa"] = holdings["saa_taa"].astype(str).str.strip().str.upper()
    holdings["Fund Code"] = holdings["Fund Code"].astype(str).str.strip()
    holdings["Fund Description"] = holdings["Fund Description"].apply(normalize_text)
    holdings = holdings[holdings["Total MV (CAD)"].fillna(0) > 0].copy()
    holdings = holdings[holdings["mandate_code"].notna()].copy()
    holdings = holdings[holdings["saa_taa"].isin(["SAA", "TAA"])].copy()
    holdings = holdings.reset_index(drop=True)

    messages = {
        "warnings": [],
        "info": [],
    }

    if holdings.empty:
        raise ValueError("No IPS holdings rows with `Total MV (CAD)` greater than 0 were found.")

    duplicate_codes = holdings["mandate_code"][holdings["mandate_code"].duplicated()].unique().tolist()
    if duplicate_codes:
        messages["warnings"].append(
            f"Duplicate mandate codes detected in holdings and will reuse the same support file: {', '.join(sorted(duplicate_codes))}."
        )

    return holdings, messages


def derive_support_code_from_fund_code(fund_code: object) -> Optional[str]:
    text = normalize_text(fund_code)
    if not text.isdigit():
        return None

    numeric = int(text)
    derived = numeric - 2000
    if 10000 <= derived <= 99999:
        return str(derived)
    return None


def build_support_candidates(fund_code: object, mandate_code: object = None) -> List[str]:
    candidates: List[str] = []
    entered = extract_mandate_code(mandate_code)
    derived = derive_support_code_from_fund_code(fund_code)

    if derived:
        candidates.append(derived)

    if entered:
        candidates.append(entered)
        if entered.isdigit():
            candidates.append(str(int(entered) + 1).zfill(len(entered)))
            if int(entered) > 0:
                candidates.append(str(int(entered) - 1).zfill(len(entered)))

    deduped: List[str] = []
    for code in candidates:
        if code and code not in deduped:
            deduped.append(code)
    return deduped


def load_draft_state() -> Tuple[pd.DataFrame, List[dict]]:
    if not DRAFT_PATH.exists():
        return DEFAULT_HOLDINGS_INPUT.copy(), []

    try:
        payload = json.loads(DRAFT_PATH.read_text())
    except Exception:
        return DEFAULT_HOLDINGS_INPUT.copy(), []

    holdings_records = payload.get("holdings", [])
    holdings_text = payload.get("holdings_text")
    support_records = payload.get("support_files", [])

    if holdings_text:
        try:
            holdings_df = parse_holdings_text(holdings_text)
        except Exception:
            holdings_df = pd.DataFrame(holdings_records) if holdings_records else DEFAULT_HOLDINGS_INPUT.copy()
    else:
        holdings_df = pd.DataFrame(holdings_records) if holdings_records else DEFAULT_HOLDINGS_INPUT.copy()
    for column in MANUAL_HOLDINGS_COLUMNS:
        if column not in holdings_df.columns:
            holdings_df[column] = DEFAULT_HOLDINGS_INPUT.iloc[0].get(column, "")
    holdings_df = holdings_df[MANUAL_HOLDINGS_COLUMNS].copy()

    saved_support_files: List[dict] = []
    for record in support_records:
        filename = record.get("filename")
        content_b64 = record.get("content_b64")
        if not filename or not content_b64:
            continue
        try:
            saved_support_files.append(
                {
                    "filename": filename,
                    "bytes": base64.b64decode(content_b64.encode("ascii")),
                }
            )
        except Exception:
            continue

    return holdings_df, saved_support_files


def save_draft_state(holdings_df: pd.DataFrame, saved_support_files: List[dict], holdings_text: str) -> None:
    payload = {
        "holdings": holdings_df.where(pd.notnull(holdings_df), None).to_dict("records"),
        "holdings_text": holdings_text,
        "support_files": [
            {
                "filename": item["filename"],
                "content_b64": base64.b64encode(item["bytes"]).decode("ascii"),
            }
            for item in saved_support_files
        ],
    }
    DRAFT_PATH.write_text(json.dumps(payload))


def clear_draft_state() -> None:
    if DRAFT_PATH.exists():
        DRAFT_PATH.unlink()


def blank_holding_row() -> dict:
    return {
        "Fund Code": "",
        "Fund Description": "",
        "Total MV (CAD)": "",
        "saa_taa": "SAA",
    }


def pad_holding_rows(rows: List[dict], minimum: int = 5) -> List[dict]:
    padded = list(rows)
    while len(padded) < minimum:
        padded.append(blank_holding_row())
    return padded


def holding_rows_to_df(rows: List[dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)
    return pd.DataFrame(rows)[MANUAL_HOLDINGS_COLUMNS].copy()


def holdings_df_to_text(df: pd.DataFrame) -> str:
    if df is None or df.empty:
        return HOLDINGS_TEXT_TEMPLATE
    export_df = df[MANUAL_HOLDINGS_COLUMNS].copy()
    return export_df.to_csv(sep="\t", index=False)


def parse_holdings_text(holdings_text: str) -> pd.DataFrame:
    text = holdings_text.strip()
    if not text:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)

    parsed = pd.read_csv(StringIO(text), sep=None, engine="python")
    normalized_columns = {col: normalize_header(col) for col in parsed.columns}

    if set(normalized_columns.values()) >= {"FUNDCODE", "FUNDDESCRIPTION", "TOTALMVCAD", "SAATAA"}:
        rename_map = {}
        for col, normalized in normalized_columns.items():
            if normalized == "FUNDCODE":
                rename_map[col] = "Fund Code"
            elif normalized == "FUNDDESCRIPTION":
                rename_map[col] = "Fund Description"
            elif normalized == "TOTALMVCAD":
                rename_map[col] = "Total MV (CAD)"
            elif normalized == "SAATAA":
                rename_map[col] = "saa_taa"
        parsed = parsed.rename(columns=rename_map)
        return parsed[MANUAL_HOLDINGS_COLUMNS].copy()

    parsed = pd.read_csv(
        StringIO(text),
        sep=None,
        engine="python",
        header=None,
        names=MANUAL_HOLDINGS_COLUMNS,
    )
    return parsed[MANUAL_HOLDINGS_COLUMNS].copy()


def parse_manual_holdings_input(input_df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    if input_df is None:
        raise ValueError("Enter at least one holdings row before running the calculation.")

    holdings = input_df.copy()
    if "mandate_code" not in holdings.columns:
        holdings["mandate_code"] = ""

    required_columns = MANUAL_HOLDINGS_COLUMNS + ["mandate_code"]
    missing_columns = [col for col in MANUAL_HOLDINGS_COLUMNS if col not in holdings.columns]
    if missing_columns:
        raise ValueError(f"Missing required holdings columns: {', '.join(missing_columns)}")

    holdings = holdings[required_columns].copy()
    holdings["Fund Code"] = holdings["Fund Code"].apply(normalize_text)
    holdings["Fund Description"] = holdings["Fund Description"].apply(normalize_text)
    holdings["Total MV (CAD)"] = coerce_number_series(holdings["Total MV (CAD)"])
    holdings["mandate_code"] = holdings["mandate_code"].apply(extract_mandate_code)
    holdings["saa_taa"] = holdings["saa_taa"].astype(str).str.strip().str.upper()

    messages = {
        "warnings": [],
        "info": [],
    }

    all_blank_mask = (
        (holdings["Fund Code"] == "")
        & (holdings["Fund Description"] == "")
        & (holdings["Total MV (CAD)"].isna())
    )
    blank_rows = int(all_blank_mask.sum())
    if blank_rows:
        messages["info"].append(f"Ignored {blank_rows} blank holdings row(s).")
    holdings = holdings[~all_blank_mask].copy()

    if holdings.empty:
        raise ValueError("Enter at least one complete holdings row before running the calculation.")

    invalid_saa_taa = holdings[~holdings["saa_taa"].isin(["SAA", "TAA"])]
    if not invalid_saa_taa.empty:
        raise ValueError("Each holdings row must have `saa_taa` set to either `SAA` or `TAA`.")

    missing_required = holdings[
        (holdings["Fund Code"] == "")
        | (holdings["Fund Description"] == "")
        | (holdings["Total MV (CAD)"].isna())
    ]
    if not missing_required.empty:
        raise ValueError(
            "Each holdings row must include Fund Code, Fund Description, Total MV (CAD), and saa_taa."
        )

    zero_mv_rows = holdings[holdings["Total MV (CAD)"].fillna(0) <= 0]
    if not zero_mv_rows.empty:
        messages["warnings"].append(
            f"Ignored {len(zero_mv_rows)} holdings row(s) with Total MV (CAD) less than or equal to 0."
        )
    holdings = holdings[holdings["Total MV (CAD)"].fillna(0) > 0].copy()

    if holdings.empty:
        raise ValueError("All entered holdings rows have Total MV (CAD) less than or equal to 0.")

    holdings["support_code"] = holdings["Fund Code"].apply(derive_support_code_from_fund_code)
    unresolved_support_rows = holdings[holdings["support_code"].isna()]
    if not unresolved_support_rows.empty:
        raise ValueError(
            "Unable to derive the expected support file code from one or more Fund Code values."
        )

    duplicate_codes = holdings["support_code"][holdings["support_code"].duplicated()].unique().tolist()
    if duplicate_codes:
        messages["warnings"].append(
            f"Duplicate support codes detected in holdings and will reuse the same support file: {', '.join(sorted(duplicate_codes))}."
        )

    holdings = holdings.reset_index(drop=True)
    return holdings, messages


def parse_manual_holdings_text(input_text: str) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    parsed_df = parse_holdings_text(input_text)
    return parse_manual_holdings_input(parsed_df)


def find_match_position(df: pd.DataFrame, target: str) -> Optional[Tuple[int, int]]:
    target_key = normalize_key(target)
    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            if normalize_key(df.iat[row_idx, col_idx]) == target_key:
                return row_idx, col_idx
    return None


def parse_support_report_date(raw_df: pd.DataFrame, anchor_row: int) -> Optional[pd.Timestamp]:
    window_start = max(anchor_row - 12, 0)
    date_pattern = re.compile(r"\b\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b", re.IGNORECASE)

    for row_idx in range(anchor_row, window_start - 1, -1):
        for value in raw_df.iloc[row_idx].tolist():
            if pd.isna(value):
                continue
            if isinstance(value, pd.Timestamp):
                return value
            text = normalize_text(value)
            if not text:
                continue
            match = date_pattern.search(text)
            if not match:
                continue
            parsed = pd.to_datetime(match.group(0), errors="coerce", dayfirst=True)
            if pd.notna(parsed):
                return parsed

    return None


def format_reporting_period(report_date: Optional[pd.Timestamp]) -> Optional[str]:
    if report_date is None or pd.isna(report_date):
        return None
    return pd.Timestamp(report_date).strftime("%B %Y")


def build_reporting_period_note(reporting_period: Optional[str]) -> Optional[str]:
    if not reporting_period:
        return None
    return f"Asset Allocation sector weight is reported as of {reporting_period}."


def parse_support_file(file_bytes: bytes, filename: str) -> Tuple[pd.DataFrame, Optional[pd.Timestamp]]:
    raw_df = read_excel_sheet(file_bytes, filename, sheet_name=0, header=None)
    port_weight_pos = find_match_position(raw_df, "Port. Weight")
    non_composite_pos = find_match_position(raw_df, "Non-Composite")

    if port_weight_pos is None or non_composite_pos is None:
        raise ValueError(
            f"`{filename}` does not contain the expected `Port. Weight` / `Non-Composite` block."
        )

    start_row, name_col = non_composite_pos
    _, weight_col = port_weight_pos
    report_date = parse_support_report_date(raw_df, start_row)
    subset = raw_df.iloc[start_row:, [name_col, weight_col]].copy()
    subset.columns = ["Component", "Port. Weight"]
    subset["Component"] = subset["Component"].apply(normalize_text)
    subset["Port. Weight"] = coerce_number_series(subset["Port. Weight"])
    subset = subset.dropna(how="all")
    subset = subset[subset["Component"] != ""]
    subset = subset[subset["Port. Weight"].notna()]
    component_key = subset["Component"].map(normalize_key)
    duplicate_adjacent = (
        component_key.eq(component_key.shift(-1))
        & subset["Port. Weight"].round(12).eq(subset["Port. Weight"].shift(-1).round(12))
    )
    subset = subset[~duplicate_adjacent]
    subset = subset.reset_index(drop=True)

    if subset.empty:
        raise ValueError(f"No support rows were parsed from `{filename}`.")

    return subset, report_date


def apply_composition_mapping(component: pd.Series) -> pd.Series:
    normalized = component.map(normalize_key)
    base = normalized.map(ASSET_CLASS_DETECTION)
    override = normalized.map(FUNDS_ALTERNATIVES_DETECTION)
    final = override.fillna(base).fillna("")
    final = final.replace(
        {
            "Private Alt": "Private Alternatives",
            "Liquid Alt": "Private Alternatives",
        }
    )
    return final


def apply_breakdown_mapping(component: pd.Series) -> pd.Series:
    normalized = component.map(normalize_key)
    result = pd.Series("", index=component.index, dtype="object")
    result.loc[normalized.isin(BREAKDOWN_ALTERNATIVES)] = "Alternatives"
    result.loc[normalized.isin(BREAKDOWN_CASH)] = "Cash"
    result.loc[normalized.isin(BREAKDOWN_OTHER)] = "Other"
    for key, label in BREAKDOWN_DIRECT.items():
        result.loc[normalized == key] = label
    return result


def build_uploaded_support_map(files: Iterable[object]) -> Tuple[Dict[str, dict], List[str]]:
    support_map: Dict[str, dict] = {}
    warnings: List[str] = []

    for uploaded in files:
        code = extract_mandate_code(uploaded.name)
        if not code:
            warnings.append(f"Ignored support file `{uploaded.name}` because no 5-digit mandate code was found in the filename.")
            continue
        if code in support_map:
            warnings.append(f"Duplicate support file uploaded for mandate code {code}. Using `{support_map[code]['filename']}` and ignoring `{uploaded.name}`.")
            continue
        support_map[code] = {
            "filename": uploaded.name,
            "bytes": uploaded.getvalue(),
        }

    return support_map, warnings


def build_saved_support_map(saved_files: List[dict]) -> Tuple[Dict[str, dict], List[str]]:
    support_map: Dict[str, dict] = {}
    warnings: List[str] = []

    for saved in saved_files:
        filename = saved["filename"]
        code = extract_mandate_code(filename)
        if not code:
            warnings.append(f"Ignored support file `{filename}` because no 5-digit code was found in the filename.")
            continue
        if code in support_map:
            warnings.append(f"Duplicate support file saved for code {code}. Using `{support_map[code]['filename']}` and ignoring `{filename}`.")
            continue
        support_map[code] = {
            "filename": filename,
            "bytes": saved["bytes"],
        }

    return support_map, warnings


def resolve_support_file(holding: pd.Series, support_map: Dict[str, dict]) -> Tuple[Optional[str], Optional[dict], List[str]]:
    candidates = build_support_candidates(holding["Fund Code"], holding.get("mandate_code"))
    for code in candidates:
        if code in support_map:
            return code, support_map[code], candidates
    return None, None, candidates


def calculate_reports(holdings: pd.DataFrame, support_map: Dict[str, dict]) -> Tuple[dict, List[str], List[str]]:
    warnings: List[str] = []
    info: List[str] = []
    blocks: List[pd.DataFrame] = []
    matched_codes: set[str] = set()
    unresolved_holdings: List[str] = []
    reporting_periods: set[str] = set()

    for _, holding in holdings.iterrows():
        code, support_file, candidates = resolve_support_file(holding, support_map)
        if support_file is None or code is None:
            descriptor = f"{holding['Fund Description']} (expected one of: {', '.join(candidates) if candidates else 'no derived support code'})"
            unresolved_holdings.append(descriptor)
            continue

        matched_codes.add(code)
        support_rows, report_date = parse_support_file(support_file["bytes"], support_file["filename"])
        reporting_period = format_reporting_period(report_date)
        if reporting_period:
            reporting_periods.add(reporting_period)
        support_rows["Fund Code"] = holding["Fund Code"]
        support_rows["Fund Description"] = holding["Fund Description"]
        support_rows["Total MV (CAD)"] = float(holding["Total MV (CAD)"])
        support_rows["mandate_code"] = holding.get("mandate_code")
        support_rows["support_code"] = code
        support_rows["saa_taa"] = holding["saa_taa"]
        support_rows["Support File"] = support_file["filename"]
        support_rows["Block Label"] = f"{holding['Fund Description']} - {code} - {holding['saa_taa']}"
        support_rows["Weighted MV (CAD)"] = support_rows["Total MV (CAD)"] * support_rows["Port. Weight"] / 100.0
        blocks.append(support_rows)

    if unresolved_holdings:
        raise ValueError(
            "Missing support files for holdings: " + "; ".join(unresolved_holdings)
        )

    extra_codes = sorted(set(support_map) - matched_codes)
    if extra_codes:
        ignored = [support_map[code]["filename"] for code in extra_codes]
        warnings.append("Ignoring extra uploaded support files not referenced in holdings: " + ", ".join(ignored))

    comp_df = pd.concat(blocks, ignore_index=True)
    comp_df["Composition Group"] = apply_composition_mapping(comp_df["Component"])
    comp_df["Breakdown Group"] = apply_breakdown_mapping(comp_df["Component"])

    portfolio_total = float(holdings["Total MV (CAD)"].sum())
    if portfolio_total <= 0:
        raise ValueError("The total holdings market value must be greater than 0.")

    composition_rows = comp_df[comp_df["Composition Group"] != ""].copy()
    breakdown_rows = comp_df[comp_df["Breakdown Group"] != ""].copy()

    composition_summary = (
        composition_rows.groupby("Composition Group", as_index=True)["Weighted MV (CAD)"]
        .sum()
        .reindex(COMPOSITION_GROUP_ORDER, fill_value=0.0)
    )
    composition_df = pd.DataFrame(
        {
            "Asset Classes": COMPOSITION_GROUP_ORDER,
            "Portfolio Market Value (CDN)": composition_summary.values,
            "Actively Managed Market Value (CDN)": composition_summary.values,
            "% of Portfolio (CDN)": composition_summary.values / portfolio_total * 100.0,
        }
    )
    composition_total_row = pd.DataFrame(
        [
            {
                "Asset Classes": "Total Market Value of Asset Classes",
                "Portfolio Market Value (CDN)": composition_df["Portfolio Market Value (CDN)"].sum(),
                "Actively Managed Market Value (CDN)": composition_df["Actively Managed Market Value (CDN)"].sum(),
                "% of Portfolio (CDN)": composition_df["% of Portfolio (CDN)"].sum(),
            }
        ]
    )
    composition_df = pd.concat([composition_df, composition_total_row], ignore_index=True)

    breakdown_pivot = (
        breakdown_rows.pivot_table(
            index="Breakdown Group",
            columns="saa_taa",
            values="Weighted MV (CAD)",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(BREAKDOWN_GROUP_ORDER, fill_value=0.0)
    )
    for column in ["SAA", "TAA"]:
        if column not in breakdown_pivot.columns:
            breakdown_pivot[column] = 0.0
    breakdown_pivot = breakdown_pivot[["SAA", "TAA"]]

    breakdown_df = pd.DataFrame(
        {
            "Actively Managed Asset Classes": BREAKDOWN_GROUP_ORDER,
            "Strategic Asset Allocation %": breakdown_pivot["SAA"].values / portfolio_total * 100.0,
            "Tactical Asset Allocation %": breakdown_pivot["TAA"].values / portfolio_total * 100.0,
        }
    )
    breakdown_df["Portfolio %"] = (
        breakdown_df["Strategic Asset Allocation %"] + breakdown_df["Tactical Asset Allocation %"]
    )
    breakdown_total_row = pd.DataFrame(
        [
            {
                "Actively Managed Asset Classes": "Total of Actively Managed Assets",
                "Strategic Asset Allocation %": breakdown_df["Strategic Asset Allocation %"].sum(),
                "Tactical Asset Allocation %": breakdown_df["Tactical Asset Allocation %"].sum(),
                "Portfolio %": breakdown_df["Portfolio %"].sum(),
            }
        ]
    )
    breakdown_df = pd.concat([breakdown_df, breakdown_total_row], ignore_index=True)

    composition_mapped_total = composition_rows["Weighted MV (CAD)"].sum()
    if abs(composition_mapped_total - portfolio_total) > 0.5:
        unmatched = portfolio_total - composition_mapped_total
        warnings.append(
            f"Composition mapping did not reconcile exactly to holdings total. Unmapped amount: ${unmatched:,.2f}."
        )

    breakdown_mapped_total = breakdown_rows["Weighted MV (CAD)"].sum()
    if abs(breakdown_mapped_total - portfolio_total) > 0.5:
        unmatched = portfolio_total - breakdown_mapped_total
        warnings.append(
            f"Breakdown mapping did not reconcile exactly to holdings total. Unmapped amount: ${unmatched:,.2f}."
        )

    comp_block_check = composition_rows.groupby("Block Label")["Port. Weight"].sum()
    off_comp = comp_block_check[comp_block_check.sub(100).abs() > 0.25]
    if not off_comp.empty:
        warnings.append(
            "Some support blocks do not sum to 100% after composition mapping: "
            + ", ".join(f"{label} ({value:.2f}%)" for label, value in off_comp.items())
        )

    break_block_check = breakdown_rows.groupby("Block Label")["Port. Weight"].sum()
    off_break = break_block_check[break_block_check.sub(100).abs() > 0.25]
    if not off_break.empty:
        warnings.append(
            "Some support blocks do not sum to 100% after breakdown mapping: "
            + ", ".join(f"{label} ({value:.2f}%)" for label, value in off_break.items())
        )

    equity_composition = float(
        composition_df.loc[composition_df["Asset Classes"] == "Equity", "Portfolio Market Value (CDN)"].iloc[0]
    )
    equity_breakdown = float(
        breakdown_pivot.loc[["International Equity", "US Equity", "Canadian Equity"]].sum().sum()
    )
    if abs(equity_composition - equity_breakdown) > 0.5:
        warnings.append(
            f"Equity reconciliation failed. Composition Equity = ${equity_composition:,.2f}, Breakdown Equity Total = ${equity_breakdown:,.2f}."
        )

    if holdings["saa_taa"].eq("SAA").all():
        info.append("This is an SAA-only portfolio. Tactical values remain at 0.00%.")
    if holdings["saa_taa"].eq("TAA").all():
        info.append("This is a TAA-only portfolio. Strategic values remain at 0.00%.")

    reporting_period = None
    if len(reporting_periods) == 1:
        reporting_period = next(iter(reporting_periods))
        info.append(f"Detected support-file reporting period: {reporting_period}.")
    elif len(reporting_periods) > 1:
        warnings.append(
            "Uploaded support files appear to span multiple reporting periods: "
            + ", ".join(sorted(reporting_periods))
            + "."
        )
    else:
        warnings.append(
            "No reporting period could be detected from the uploaded support files."
        )

    return {
        "holdings": holdings.copy(),
        "comp_detail": comp_df,
        "composition": composition_df,
        "breakdown": breakdown_df,
        "portfolio_total": portfolio_total,
        "reporting_period": reporting_period,
    }, warnings, info


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def format_percent(value: float) -> str:
    return f"{value:,.2f}%"


def build_composition_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for col in ["Portfolio Market Value (CDN)", "Actively Managed Market Value (CDN)"]:
        display[col] = display[col].map(format_currency)
    display["% of Portfolio (CDN)"] = display["% of Portfolio (CDN)"].map(format_percent)
    return display


def build_breakdown_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        display[col] = display[col].map(format_percent)
    return display


def build_composition_chart(df: pd.DataFrame) -> go.Figure:
    chart_df = df[df["Asset Classes"] != "Total Market Value of Asset Classes"].copy()
    chart_df = chart_df[chart_df["Portfolio Market Value (CDN)"] > 0].copy()
    fig = px.pie(
        chart_df,
        names="Asset Classes",
        values="Portfolio Market Value (CDN)",
        hole=0.6,
        color="Asset Classes",
        color_discrete_map={
            "Income": PRIMARY_BLUE,
            "Equity": "#4B4B4B",
            "Cash": "#D7D1C3",
            "Other": "#A7A7A7",
            "Private Alternatives": "#C7B168",
        },
    )
    fig.update_traces(
        textposition="outside",
        texttemplate="%{label} %{percent:.2%}",
        textfont=dict(color="#111827", size=16),
        marker=dict(line=dict(color="white", width=2)),
        sort=False,
        direction="clockwise",
    )
    fig.update_layout(
        margin=dict(l=90, r=120, t=30, b=40),
        showlegend=False,
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111827", size=16),
        uniformtext=dict(minsize=14, mode="hide"),
    )
    fig.update_traces(automargin=True)
    return fig


def build_breakdown_chart(df: pd.DataFrame) -> go.Figure:
    chart_df = df[df["Actively Managed Asset Classes"] != "Total of Actively Managed Assets"].copy()
    chart_df = chart_df.iloc[::-1].reset_index(drop=True)
    long_df = chart_df.melt(
        id_vars="Actively Managed Asset Classes",
        value_vars=["Strategic Asset Allocation %", "Tactical Asset Allocation %"],
        var_name="Allocation Type",
        value_name="Percent",
    )
    fig = px.bar(
        long_df,
        x="Percent",
        y="Actively Managed Asset Classes",
        color="Allocation Type",
        orientation="h",
        barmode="relative",
        text="Percent",
        color_discrete_map={
            "Strategic Asset Allocation %": PRIMARY_BLUE,
            "Tactical Asset Allocation %": SECONDARY_GRAY,
        },
        category_orders={
            "Actively Managed Asset Classes": chart_df["Actively Managed Asset Classes"].tolist(),
        },
    )
    fig.update_traces(texttemplate="%{text:.2f}%", textposition="inside", insidetextanchor="middle")
    fig.update_traces(
        textfont=dict(color="white", size=17),
        selector=dict(name="Strategic Asset Allocation %"),
    )
    fig.update_traces(
        textfont=dict(color="#1F2937", size=17),
        selector=dict(name="Tactical Asset Allocation %"),
    )
    totals = chart_df["Portfolio %"].tolist()
    labels = chart_df["Actively Managed Asset Classes"].tolist()
    max_total = max(totals) if totals else 0.0
    x_axis_max = max(max_total * 1.18, 10.0)
    fig.update_layout(
        margin=dict(l=180, r=120, t=30, b=40),
        legend=dict(
            title="",
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="left",
            x=0,
            font=dict(color="#111827", size=16),
        ),
        xaxis_title="Portfolio %",
        yaxis_title="",
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111827", size=16),
        xaxis=dict(
            range=[0, x_axis_max],
            automargin=True,
            tickfont=dict(color="#111827", size=16),
            title_font=dict(color="#111827", size=16),
            gridcolor="#E5E7EB",
            zerolinecolor="#D1D5DB",
        ),
        yaxis=dict(
            automargin=True,
            tickfont=dict(color="#111827", size=18),
            title_font=dict(color="#111827", size=16),
        ),
        uniformtext=dict(minsize=12, mode="hide"),
    )
    fig.add_trace(
        go.Scatter(
            x=totals,
            y=labels,
            mode="text",
            text=[f"{value:.2f}%" for value in totals],
            textposition="middle right",
            textfont=dict(color="#111827", size=18),
            showlegend=False,
            hoverinfo="skip",
            cliponaxis=False,
        )
    )
    return fig


def maybe_render_figure_png(fig: go.Figure, width: int, height: int) -> Optional[bytes]:
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        return None


def autofit_columns(worksheet) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def write_table(
    worksheet,
    start_row: int,
    dataframe: pd.DataFrame,
    currency_columns: Optional[Iterable[str]] = None,
    percent_columns: Optional[Iterable[str]] = None,
    title: Optional[str] = None,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_font = Font(bold=True)
    row = start_row
    if title:
        worksheet.cell(row=row, column=1, value=title).font = Font(size=14, bold=True)
        row += 2

    for col_idx, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=row, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    currency_columns = set(currency_columns or [])
    percent_columns = set(percent_columns or [])

    for record_idx, (_, record) in enumerate(dataframe.iterrows(), start=0):
        is_total = record_idx == len(dataframe) - 1
        for col_idx, column_name in enumerate(dataframe.columns, start=1):
            cell = worksheet.cell(row=row, column=col_idx, value=record[column_name])
            if column_name in currency_columns:
                cell.number_format = '$#,##0.00'
            elif column_name in percent_columns:
                cell.number_format = '0.00%'
            if is_total:
                cell.fill = total_fill
                cell.font = total_font
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
        row += 1

    autofit_columns(worksheet)
    worksheet.freeze_panes = worksheet["A3"] if title else worksheet["A2"]
    return row


def build_excel_report(
    holdings_df: pd.DataFrame,
    composition_df: pd.DataFrame,
    breakdown_df: pd.DataFrame,
    composition_fig: go.Figure,
    breakdown_fig: go.Figure,
    reporting_period: Optional[str],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError(
            "Excel export requires `openpyxl`. Install it with `pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    holdings_ws = workbook.create_sheet("Holdings")
    holdings_export = holdings_df.copy()
    holdings_export["Total MV (CAD)"] = holdings_export["Total MV (CAD)"].astype(float)
    write_table(
        holdings_ws,
        start_row=1,
        dataframe=holdings_export,
        currency_columns=["Total MV (CAD)"],
        title="Holdings",
    )

    composition_ws = workbook.create_sheet("Composition")
    composition_export = composition_df.copy()
    composition_export["% of Portfolio (CDN)"] = composition_export["% of Portfolio (CDN)"] / 100.0
    comp_end_row = write_table(
        composition_ws,
        start_row=1,
        dataframe=composition_export,
        currency_columns=["Portfolio Market Value (CDN)", "Actively Managed Market Value (CDN)"],
        percent_columns=["% of Portfolio (CDN)"],
        title="Portfolio Composition",
    )

    breakdown_ws = workbook.create_sheet("Breakdown")
    breakdown_export = breakdown_df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        breakdown_export[col] = breakdown_export[col] / 100.0
    break_end_row = write_table(
        breakdown_ws,
        start_row=1,
        dataframe=breakdown_export,
        percent_columns=["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"],
        title="Portfolio Breakdown",
    )
    next_note_row = break_end_row + 2
    reporting_period_note = build_reporting_period_note(reporting_period)
    if reporting_period_note:
        note_one = breakdown_ws.cell(
            row=next_note_row,
            column=1,
            value=reporting_period_note,
        )
        note_one.font = Font(italic=True)
        next_note_row += 1
    note_two = breakdown_ws.cell(
        row=next_note_row,
        column=1,
        value="The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.",
    )
    note_two.font = Font(italic=True)

    if PILImage is not None:
        comp_png = maybe_render_figure_png(composition_fig, width=1300, height=900)
        break_png = maybe_render_figure_png(breakdown_fig, width=1500, height=900)

        if comp_png:
            image = XLImage(PILImage.open(BytesIO(comp_png)))
            image.width = 820
            image.height = 568
            composition_ws.add_image(image, f"A{comp_end_row + 2}")

        if break_png:
            image = XLImage(PILImage.open(BytesIO(break_png)))
            image.width = 940
            image.height = 564
            breakdown_ws.add_image(image, f"A{break_end_row + 5}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def show_validation_messages(warnings: Iterable[str], info: Iterable[str]) -> None:
    for message in warnings:
        st.warning(message)
    for message in info:
        st.info(message)


st.markdown(
    """
    <style>
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .app-subtitle {
        color: #4b5563;
        margin-top: -0.75rem;
        margin-bottom: 1rem;
        font-size: 1rem;
    }
    .section-card {
        padding: 1rem 1.1rem;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Portfolio Composition Calculator")
st.markdown(
    '<div class="app-subtitle">Streamlit calculator for Portfolio Composition and Portfolio Breakdown reporting.</div>',
    unsafe_allow_html=True,
)
st.warning(
    "Important: Ensure you upload the correct support files matching the funds in your holdings. Incorrect files will produce wrong results."
)

with st.expander("Portfolio Composition Calculator Manual"):
    st.markdown(MANUAL_TEXT)

if "draft_initialized" not in st.session_state:
    draft_holdings, draft_support_files = load_draft_state()
    st.session_state["holdings_rows"] = pad_holding_rows(draft_holdings.to_dict("records"))
    st.session_state["holdings_paste_text"] = holdings_df_to_text(draft_holdings)
    st.session_state["saved_support_files"] = draft_support_files
    st.session_state["draft_initialized"] = True

with st.sidebar:
    st.header("Support Files")
    support_files = st.file_uploader(
        "Upload Support Files (exactly the files needed – typically 5 XLSX)",
        type=["xlsx", "xls", "xlsm"],
        accept_multiple_files=True,
        help="Upload only the mandate support files referenced by the holdings file.",
    )
    if st.button("Reset Saved Draft", width="stretch"):
        clear_draft_state()
        st.session_state["holdings_rows"] = pad_holding_rows([])
        st.session_state["holdings_paste_text"] = HOLDINGS_TEXT_TEMPLATE
        st.session_state["saved_support_files"] = []
        for key in list(st.session_state.keys()):
            if key.startswith("holding_") or key == "holdings_paste_area":
                del st.session_state[key]
        st.rerun()

    st.caption("Entered holdings and uploaded support files are saved locally so a refresh does not clear your work.")
    run_calculation = st.button(
        "Run Calculation",
        type="primary",
        width="stretch",
        disabled=not support_files and not st.session_state.get("saved_support_files"),
    )

st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.subheader("Enter IPS Holdings")
st.write("Five rows are shown by default. Add more rows if needed. Support-file matching is derived automatically from the fund code.")

if "holdings_rows" not in st.session_state:
    st.session_state["holdings_rows"] = pad_holding_rows([])
if "holdings_paste_text" not in st.session_state:
    st.session_state["holdings_paste_text"] = HOLDINGS_TEXT_TEMPLATE

with st.expander("Paste From Excel"):
    st.write("Paste tab-separated rows from Excel. Include columns: `Fund Code`, `Fund Description`, `Total MV (CAD)`, `saa_taa`.")
    pasted_text = st.text_area(
        "Bulk Paste",
        value=st.session_state["holdings_paste_text"],
        height=180,
        key="holdings_paste_area",
        label_visibility="collapsed",
    )
    paste_col1, paste_col2, _ = st.columns([1.2, 1.2, 3.6])
    if paste_col1.button("Replace Rows", width="stretch"):
        imported_df = parse_holdings_text(pasted_text)
        st.session_state["holdings_rows"] = pad_holding_rows(imported_df.to_dict("records"))
        st.session_state["holdings_paste_text"] = pasted_text
        st.rerun()
    if paste_col2.button("Append Rows", width="stretch"):
        imported_df = parse_holdings_text(pasted_text)
        current_df = holding_rows_to_df(st.session_state["holdings_rows"])
        combined_df = pd.concat([current_df, imported_df], ignore_index=True)
        st.session_state["holdings_rows"] = pad_holding_rows(combined_df.to_dict("records"))
        st.session_state["holdings_paste_text"] = pasted_text
        st.rerun()

control_col1, control_col2, _ = st.columns([1, 1, 4])
if control_col1.button("Add Row", width="stretch"):
    st.session_state["holdings_rows"].append(blank_holding_row())
    st.rerun()
if control_col2.button("Add 5 Rows", width="stretch"):
    st.session_state["holdings_rows"].extend([blank_holding_row() for _ in range(5)])
    st.rerun()

header_cols = st.columns([1.1, 3.6, 1.3, 1.0])
header_cols[0].markdown("**Fund Code**")
header_cols[1].markdown("**Fund Description**")
header_cols[2].markdown("**Total MV (CAD)**")
header_cols[3].markdown("**SAA/TAA**")

rendered_rows: List[dict] = []
for idx, row in enumerate(st.session_state["holdings_rows"]):
    cols = st.columns([1.1, 3.6, 1.3, 1.0])
    fund_code = cols[0].text_input(
        f"Fund Code {idx + 1}",
        value=row.get("Fund Code", ""),
        label_visibility="collapsed",
        key=f"holding_fund_code_{idx}",
    )
    fund_description = cols[1].text_input(
        f"Fund Description {idx + 1}",
        value=row.get("Fund Description", ""),
        label_visibility="collapsed",
        key=f"holding_fund_description_{idx}",
    )
    total_mv = cols[2].text_input(
        f"Total MV (CAD) {idx + 1}",
        value="" if row.get("Total MV (CAD)", "") is None else str(row.get("Total MV (CAD)", "")),
        label_visibility="collapsed",
        key=f"holding_total_mv_{idx}",
    )
    saa_taa = cols[3].selectbox(
        f"SAA/TAA {idx + 1}",
        options=["SAA", "TAA"],
        index=0 if row.get("saa_taa", "SAA") != "TAA" else 1,
        label_visibility="collapsed",
        key=f"holding_saa_taa_{idx}",
    )
    rendered_rows.append(
        {
            "Fund Code": fund_code,
            "Fund Description": fund_description,
            "Total MV (CAD)": total_mv,
            "saa_taa": saa_taa,
        }
    )

st.session_state["holdings_rows"] = rendered_rows
preview_df = holding_rows_to_df(rendered_rows)
st.session_state["holdings_paste_text"] = holdings_df_to_text(preview_df)
if not preview_df.empty:
    st.caption("Preview")
    st.dataframe(preview_df, width="stretch", hide_index=True)
st.markdown("</div>", unsafe_allow_html=True)

if support_files:
    st.session_state["saved_support_files"] = [
        {
            "filename": uploaded.name,
            "bytes": uploaded.getvalue(),
        }
        for uploaded in support_files
    ]

save_draft_state(
    preview_df,
    st.session_state.get("saved_support_files", []),
    st.session_state.get("holdings_paste_text", HOLDINGS_TEXT_TEMPLATE),
)

active_support_files = st.session_state.get("saved_support_files", [])

if active_support_files:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Support File Review")
    for saved in active_support_files:
        st.write(f"- `{saved['filename']}`")
    st.markdown("</div>", unsafe_allow_html=True)
elif not run_calculation:
    st.info("Enter the IPS holdings above, upload the matching support files, then click `Run Calculation`.")

if run_calculation:
    progress = st.progress(0)
    status = st.empty()

    try:
        status.info("Validating entered holdings...")
        holdings_df, holdings_messages = parse_manual_holdings_input(preview_df)
        progress.progress(20)

        status.info("Indexing uploaded support files...")
        support_map, upload_warnings = build_saved_support_map(active_support_files)
        progress.progress(35)

        status.info("Parsing support files and calculating portfolio composition...")
        results, calc_warnings, calc_info = calculate_reports(holdings_df, support_map)
        progress.progress(70)

        composition_fig = build_composition_chart(results["composition"])
        breakdown_fig = build_breakdown_chart(results["breakdown"])
        progress.progress(85)

        status.info("Building Excel download...")
        excel_bytes = build_excel_report(
            results["holdings"],
            results["composition"],
            results["breakdown"],
            composition_fig,
            breakdown_fig,
            results.get("reporting_period"),
        )
        progress.progress(100)
        status.success("Calculation completed.")

        all_warnings = holdings_messages["warnings"] + upload_warnings + calc_warnings
        all_info = holdings_messages["info"] + calc_info
        show_validation_messages(all_warnings, all_info)

        summary_col1, summary_col2, summary_col3 = st.columns(3)
        summary_col1.metric("Holdings Rows Used", len(results["holdings"]))
        summary_col2.metric("Support Files Matched", len(set(results["comp_detail"]["support_code"])))
        summary_col3.metric("Total Market Value", format_currency(results["portfolio_total"]))

        composition_tab, breakdown_tab = st.tabs(["Portfolio Composition", "Portfolio Breakdown"])

        with composition_tab:
            st.dataframe(
                build_composition_display(results["composition"]),
                width="stretch",
                hide_index=True,
            )
            st.plotly_chart(composition_fig, use_container_width=True)

        with breakdown_tab:
            st.dataframe(
                build_breakdown_display(results["breakdown"]),
                width="stretch",
                hide_index=True,
            )
            st.plotly_chart(breakdown_fig, use_container_width=True)
            reporting_period_note = build_reporting_period_note(results.get("reporting_period"))
            if reporting_period_note:
                st.caption(reporting_period_note)
            st.caption("The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.")

        st.download_button(
            "Download Full Report as Excel",
            data=excel_bytes,
            file_name="portfolio_composition_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    except Exception as exc:
        progress.empty()
        status.empty()
        st.error(str(exc))

st.markdown("---")
st.caption("Version 1.0.0")
